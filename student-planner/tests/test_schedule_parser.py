from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

import openpyxl

from app.services.schedule_parser import RawCourse, parse_excel_schedule

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "sample_schedule.xlsx"


def test_parse_returns_list_of_raw_courses() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    assert isinstance(courses, list)
    assert len(courses) > 0
    assert all(isinstance(course, RawCourse) for course in courses)


def test_parse_extracts_course_name() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    names = {course.name for course in courses}
    assert "高等数学" in names
    assert "线性代数" in names
    assert "大学英语" in names


def test_parse_extracts_weekday() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == "高等数学")
    assert gaoshu.weekday == 1


def test_parse_extracts_period() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == "高等数学")
    assert gaoshu.period == "1-2"


def test_parse_extracts_teacher() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == "高等数学")
    assert gaoshu.teacher == "张老师"


def test_parse_extracts_location() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == "高等数学")
    assert gaoshu.location == "教学楼A301"


def test_parse_extracts_weeks() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gaoshu = next(course for course in courses if course.name == "高等数学")
    assert gaoshu.week_start == 1
    assert gaoshu.week_end == 16


def test_parse_handles_missing_teacher() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    tiyu = next(course for course in courses if course.name == "体育")
    assert tiyu.teacher is None


def test_parse_handles_custom_week_range() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    gailv = next(course for course in courses if course.name == "概率论")
    assert gailv.week_start == 3
    assert gailv.week_end == 16


def test_parse_total_course_count() -> None:
    courses = parse_excel_schedule(FIXTURE_PATH)
    assert len(courses) == 6


def test_parse_falls_back_to_xls_parser_when_openpyxl_rejects_stream(
    monkeypatch,
) -> None:
    expected = [
        RawCourse(
            name="高等数学",
            teacher="张老师",
            location="教学楼A301",
            weekday=1,
            period="1-2",
            week_start=1,
            week_end=16,
        )
    ]

    def fake_parse_xlsx_workbook(_file_path):
        raise BadZipFile("File is not a zip file")

    def fake_parse_xls_bytes(file_bytes: bytes):
        assert file_bytes == b"legacy-xls-bytes"
        return expected

    monkeypatch.setattr("app.services.schedule_parser._parse_xlsx_workbook", fake_parse_xlsx_workbook)
    monkeypatch.setattr("app.services.schedule_parser._parse_xls_bytes", fake_parse_xls_bytes)

    courses = parse_excel_schedule(BytesIO(b"legacy-xls-bytes"))
    assert courses == expected


def test_parse_detects_header_not_in_first_row() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "课程表"
    sheet["A1"] = "2025-2026 学年第一学期课程表"
    sheet["A3"] = "节次"
    sheet["B3"] = "周一"
    sheet["C3"] = "周二"
    sheet["A4"] = "第1-2节"
    sheet["B4"] = "高等数学\n张老师\n教学楼A301\n1-16周"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)
    assert len(courses) == 1
    assert courses[0].name == "高等数学"
    assert courses[0].weekday == 1
    assert courses[0].period == "1-2"


def test_parse_uses_sheet_with_most_courses() -> None:
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "说明"
    first["A1"] = "这是说明页，不是课表"

    second = workbook.create_sheet(title="第2页课表")
    second["A1"] = "节次"
    second["B1"] = "周三"
    second["A2"] = "3-4"
    second["B2"] = "大学英语\n李老师\n教学楼B201\n1-16周"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)
    assert len(courses) == 1
    assert courses[0].name == "大学英语"
    assert courses[0].weekday == 3


def test_parse_cell_with_blank_line_keeps_single_course_block() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "course-table"
    sheet["A1"] = "period"
    sheet["B1"] = "friday"
    sheet["A2"] = "7-8"
    sheet["B2"] = "体育\n\n操场\n1-16周"

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)

    courses = parse_excel_schedule(stream)
    assert len(courses) == 1
    assert courses[0].name == "体育"
    assert courses[0].location == "操场"
    assert courses[0].period == "7-8"
