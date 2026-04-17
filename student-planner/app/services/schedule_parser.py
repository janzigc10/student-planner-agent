"""Parse Excel course schedules into structured raw course records."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException


@dataclass
class RawCourse:
    name: str
    teacher: str | None
    location: str | None
    weekday: int
    period: str
    week_start: int
    week_end: int


_WEEKDAY_MAP: dict[str, int] = {
    "周一": 1,
    "星期一": 1,
    "Monday": 1,
    "Mon": 1,
    "周二": 2,
    "星期二": 2,
    "Tuesday": 2,
    "Tue": 2,
    "周三": 3,
    "星期三": 3,
    "Wednesday": 3,
    "Wed": 3,
    "周四": 4,
    "星期四": 4,
    "Thursday": 4,
    "Thu": 4,
    "周五": 5,
    "星期五": 5,
    "Friday": 5,
    "Fri": 5,
    "周六": 6,
    "星期六": 6,
    "Saturday": 6,
    "Sat": 6,
    "周日": 7,
    "星期日": 7,
    "星期天": 7,
    "周天": 7,
    "Sunday": 7,
    "Sun": 7,
}

_WEEK_RANGE_RE = re.compile(r"(\d+)\s*[-~～—–]\s*(\d+)\s*周")
_WEEK_EXPR_RE = re.compile(r"([0-9,，\s\-~～—–]+)\s*(?:\(\s*\[?周\]?\s*\)|\[\s*周\s*\]|周)")
_PERIOD_RE = re.compile(r"(?<![\d:])(?:第\s*)?(\d{1,2})\s*[-~～—–]\s*(\d{1,2})(?:\s*节)?(?![\d:])")
_BRACKET_PERIOD_RE = re.compile(r"\[(\d{1,2})\s*[-~～—–]\s*(\d{1,2})\s*节?\]")
_INLINE_PERIOD_RE = re.compile(r"(?:第\s*)?(\d{1,2})\s*[-~～—–]\s*(\d{1,2})\s*节")
_TEACHER_RE = re.compile(r"老师|教授|讲师|teacher|lecturer|prof", re.IGNORECASE)
_MAX_HEADER_SCAN_ROWS = 12
_MAX_PERIOD_SCAN_ROWS = 24


def parse_excel_schedule(file_path: Path | str | BinaryIO) -> list[RawCourse]:
    excel_bytes = _read_excel_bytes(file_path)

    if _is_xls_path(file_path):
        if excel_bytes is None:
            raise ValueError("Cannot read xls file.")
        return _parse_xls_bytes(excel_bytes)

    source = BytesIO(excel_bytes) if excel_bytes is not None else file_path
    try:
        return _parse_xlsx_workbook(source)
    except (BadZipFile, InvalidFileException, ValueError, OSError):
        if excel_bytes is None:
            raise
        return _parse_xls_bytes(excel_bytes)


def _parse_xlsx_workbook(file_path: Path | str | BinaryIO) -> list[RawCourse]:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    courses = _parse_xlsx_sheets(workbook)
    workbook.close()

    if courses:
        return courses

    # 部分文件依赖公式结果，data_only=True 可能拿到空值；回退读取公式文本
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    courses = _parse_xlsx_sheets(workbook)
    workbook.close()
    return courses


def _parse_xls_bytes(file_bytes: bytes) -> list[RawCourse]:
    try:
        import xlrd  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover
        raise ValueError("xlrd dependency is required to parse xls files.") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes)
    except Exception as exc:
        raise ValueError("Invalid xls workbook.") from exc

    return _parse_xls_sheets(workbook)


def _parse_xlsx_sheets(workbook: openpyxl.Workbook) -> list[RawCourse]:
    best_courses: list[RawCourse] = []
    for worksheet in workbook.worksheets:
        table = _worksheet_to_table(worksheet)
        courses = _extract_courses_from_table(table)
        if len(courses) > len(best_courses):
            best_courses = courses
    return best_courses


def _worksheet_to_table(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> list[list[str]]:
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    table: list[list[str]] = [["" for _ in range(max_col)] for _ in range(max_row)]

    for r_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            table[r_idx - 1][c_idx - 1] = _cell_to_text(value)

    # 将合并单元格左上角值填充到其余区域，避免信息丢失
    for merged in worksheet.merged_cells.ranges:
        min_col, min_row, max_col_m, max_row_m = merged.bounds
        anchor = table[min_row - 1][min_col - 1] if min_row - 1 < len(table) and min_col - 1 < len(table[0]) else ""
        if not anchor:
            continue
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                if not table[r - 1][c - 1]:
                    table[r - 1][c - 1] = anchor

    return table


def _parse_xls_sheets(workbook) -> list[RawCourse]:
    best_courses: list[RawCourse] = []
    for sheet_idx in range(workbook.nsheets):
        worksheet = workbook.sheet_by_index(sheet_idx)
        table: list[list[str]] = []
        for row_idx in range(worksheet.nrows):
            row: list[str] = []
            for col_idx in range(worksheet.ncols):
                row.append(_cell_to_text(worksheet.cell_value(row_idx, col_idx)))
            table.append(row)
        courses = _extract_courses_from_table(table)
        if len(courses) > len(best_courses):
            best_courses = courses
    return best_courses


def _extract_courses_from_table(table: list[list[str]]) -> list[RawCourse]:
    if not table:
        return []

    width = max((len(row) for row in table), default=0)
    if width == 0:
        return []

    normalized = [row + [""] * (width - len(row)) for row in table]
    candidates = _header_candidates(normalized)
    if not candidates:
        return []

    best_courses: list[RawCourse] = []
    for header_row, weekday_cols in candidates:
        period_col = _detect_period_column(normalized, header_row)
        if period_col is None:
            continue
        courses = _parse_courses_with_layout(normalized, header_row, period_col, weekday_cols)
        if len(courses) > len(best_courses):
            best_courses = courses

    return _deduplicate_courses(best_courses)


def _header_candidates(table: list[list[str]]) -> list[tuple[int, dict[int, int]]]:
    limit = min(len(table), _MAX_HEADER_SCAN_ROWS)
    candidates: list[tuple[int, dict[int, int]]] = []
    for row_idx in range(limit):
        mapping: dict[int, int] = {}
        for col_idx, text in enumerate(table[row_idx]):
            weekday = _match_weekday(text)
            if weekday is not None:
                mapping[col_idx] = weekday
        if mapping:
            candidates.append((row_idx, mapping))

    candidates.sort(key=lambda item: len(item[1]), reverse=True)
    return candidates


def _detect_period_column(table: list[list[str]], header_row: int) -> int | None:
    if header_row + 1 >= len(table):
        return None

    width = len(table[0])
    best_col: int | None = None
    best_score = 0
    data_end = min(len(table), header_row + 1 + _MAX_PERIOD_SCAN_ROWS)
    for col_idx in range(width):
        score = 0
        for row_idx in range(header_row + 1, data_end):
            if _looks_like_period_cell(table[row_idx][col_idx]):
                score += 1
        if score > best_score:
            best_score = score
            best_col = col_idx

    if best_score == 0:
        return None
    return best_col


def _parse_courses_with_layout(
    table: list[list[str]],
    header_row: int,
    period_col: int,
    weekday_cols: dict[int, int],
) -> list[RawCourse]:
    courses: list[RawCourse] = []
    for row_idx in range(header_row + 1, len(table)):
        period = _extract_period(table[row_idx][period_col])
        if period is None:
            continue

        for col_idx, weekday in weekday_cols.items():
            if col_idx == period_col:
                continue
            cell_value = table[row_idx][col_idx]
            if not cell_value.strip():
                continue
            courses.extend(_parse_cell(cell_value, weekday, period))

    return courses


def _is_xls_path(file_path: Path | str | BinaryIO) -> bool:
    if isinstance(file_path, Path):
        return file_path.suffix.lower() == ".xls"
    if isinstance(file_path, str):
        return Path(file_path).suffix.lower() == ".xls"
    return False


def _read_excel_bytes(file_path: Path | str | BinaryIO) -> bytes | None:
    if isinstance(file_path, Path):
        return file_path.read_bytes()
    if isinstance(file_path, str):
        return Path(file_path).read_bytes()

    if hasattr(file_path, "read"):
        try:
            file_path.seek(0)
        except Exception:
            return file_path.read()
        return file_path.read()

    return None


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _match_weekday(text: str) -> int | None:
    lowered = text.lower()
    for label, weekday in _WEEKDAY_MAP.items():
        if label.lower() in lowered:
            return weekday
    return None


def _extract_period(text: str) -> str | None:
    normalized = text.replace("〞", "-").replace("每", "-")
    match = _PERIOD_RE.search(normalized)
    if match is not None:
        return f"{int(match.group(1))}-{int(match.group(2))}"

    chinese_match = re.search(r"第([一二三四五六七八九十零〇]+)节", normalized)
    if chinese_match:
        numbers = _extract_chinese_period_numbers(chinese_match.group(1))
        if numbers:
            return f"{numbers[0]}-{numbers[-1]}"

    return None


def _extract_chinese_period_numbers(text: str) -> list[int]:
    token_map = {
        "一": 1,
        "二": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
        "十一": 11,
        "十二": 12,
        "十三": 13,
        "十四": 14,
        "十五": 15,
        "十六": 16,
        "十七": 17,
        "十八": 18,
        "十九": 19,
        "二十": 20,
    }
    tokens = re.findall(r"十一|十二|十三|十四|十五|十六|十七|十八|十九|二十|十|[一二三四五六七八九]", text)
    return [token_map[token] for token in tokens if token in token_map]


def _looks_like_period_cell(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    if len(stripped) > 20:
        return False
    period = _extract_period(stripped)
    if period is None:
        return False
    if any(marker in stripped for marker in ("节", "第", "上午", "下午", "晚上")):
        return True
    return bool(re.fullmatch(r"\d{1,2}\s*[-~～—–]\s*\d{1,2}", stripped))


def _parse_cell(text: str, weekday: int, period: str) -> list[RawCourse]:
    raw_blocks = [block.strip() for block in re.split(r"\n\s*\n+", text) if block.strip()]
    blocks = _merge_fragmented_blocks(raw_blocks)
    if not blocks:
        return []

    courses: list[RawCourse] = []
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue

        name = lines[0]
        teacher: str | None = None
        location: str | None = None
        week_start = 1
        week_end = 16
        resolved_period = period

        week_line_idx: int | None = None
        for idx, line in enumerate(lines[1:], start=1):
            parsed_week = _parse_week_range(line)
            if parsed_week is not None:
                week_start, week_end = parsed_week
                week_line_idx = idx
                parsed_period = _parse_period_from_text(line)
                if parsed_period is not None:
                    resolved_period = parsed_period
                break

        for idx, line in enumerate(lines[1:], start=1):
            if idx == week_line_idx:
                continue
            if teacher is None and (_TEACHER_RE.search(line) or _looks_like_teacher(line, idx)):
                teacher = line
                continue
            if location is None and _looks_like_location(line):
                location = line

        if location is None:
            for idx in range(len(lines) - 1, 0, -1):
                if idx != week_line_idx:
                    location = lines[idx]
                    break

        courses.append(
            RawCourse(
                name=name,
                teacher=teacher,
                location=location,
                weekday=weekday,
                period=resolved_period,
                week_start=week_start,
                week_end=week_end,
            )
        )

    return courses


def _merge_fragmented_blocks(blocks: list[str]) -> list[str]:
    merged: list[str] = []
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue

        if merged:
            previous_lines = [line.strip() for line in merged[-1].splitlines() if line.strip()]
            if _should_merge_with_previous_block(previous_lines, lines):
                merged[-1] = f"{merged[-1]}\n{block}"
                continue

        merged.append(block)

    return merged


def _should_merge_with_previous_block(previous_lines: list[str], current_lines: list[str]) -> bool:
    if not previous_lines or not current_lines:
        return False

    if any(_parse_week_range(line) is not None for line in previous_lines):
        return False

    if len(previous_lines) != 1:
        return False

    first_current_line = current_lines[0]
    continuation_hint = (
        _parse_week_range(first_current_line) is not None
        or bool(_TEACHER_RE.search(first_current_line))
        or _looks_like_location(first_current_line)
        or bool(re.search(r"(场|馆|中心|校区|机房)$", first_current_line))
    )
    return continuation_hint


def _parse_week_range(text: str) -> tuple[int, int] | None:
    week_match = _WEEK_RANGE_RE.search(text)
    if week_match:
        return int(week_match.group(1)), int(week_match.group(2))

    expr_match = _WEEK_EXPR_RE.search(text)
    if expr_match is None:
        return None

    expr = expr_match.group(1).replace("，", ",")
    numbers: list[int] = []
    for token in [part.strip() for part in expr.split(",") if part.strip()]:
        range_match = re.match(r"^(\d+)\s*[-~～—–]\s*(\d+)$", token)
        if range_match:
            start = int(range_match.group(1))
            end = int(range_match.group(2))
            if start <= end:
                numbers.extend(range(start, end + 1))
            else:
                numbers.extend(range(end, start + 1))
            continue
        if token.isdigit():
            numbers.append(int(token))

    if not numbers:
        return None
    return min(numbers), max(numbers)


def _parse_period_from_text(text: str) -> str | None:
    bracket_match = _BRACKET_PERIOD_RE.search(text)
    if bracket_match:
        return f"{int(bracket_match.group(1))}-{int(bracket_match.group(2))}"

    inline_match = _INLINE_PERIOD_RE.search(text)
    if inline_match:
        return f"{int(inline_match.group(1))}-{int(inline_match.group(2))}"
    return None


def _looks_like_teacher(line: str, idx: int) -> bool:
    if idx != 1:
        return False
    if _parse_week_range(line) is not None:
        return False
    if _looks_like_location(line):
        return False
    return bool(re.match(r"^[\u4e00-\u9fff·]{2,10}$", line))


def _looks_like_location(line: str) -> bool:
    lowered = line.lower()
    if any(keyword in line for keyword in ("楼", "室", "教", "实验", "馆", "场", "机房", "中心", "校区")):
        return True
    if "-" in line and any(char.isdigit() for char in line):
        return True
    return any(keyword in lowered for keyword in ("room", "building", "lab", "campus"))


def _deduplicate_courses(courses: list[RawCourse]) -> list[RawCourse]:
    seen: set[tuple[object, ...]] = set()
    deduped: list[RawCourse] = []
    for c in courses:
        key = (c.name, c.teacher, c.location, c.weekday, c.period, c.week_start, c.week_end)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(c)
    return deduped
